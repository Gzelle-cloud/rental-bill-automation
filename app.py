import os
if os.path.exists('.env'):
    from dotenv import load_dotenv
    load_dotenv()
import json
import base64
import anthropic
from flask import Flask, request, jsonify, send_file, render_template
from pypdf import PdfReader
from openpyxl import load_workbook
import tempfile
import shutil

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

XLSX_DEFAULT = os.path.join(os.path.dirname(__file__), 'kvitanciya.xlsx')

# Mapping: service name (normalized) -> Excel row numbers
# Sections: volumes (rows 5-21), corrections (rows 36-44, 68-73)
ROW_MAP = {
    # 1.1 Объём жилищных услуг (E col = col 5)
    'взнос на капитальный ремонт':        {'vol': 5,  'corr': 36},
    'водоотведение одн':                  {'vol': 6,  'corr': 37},
    'горячее в/с (носитель) одн':         {'vol': 7,  'corr': 38},
    'горячее в/с (энергия) одн':          {'vol': 8,  'corr': 39},
    'содержание жилого помещения':        {'vol': 9,  'corr': 40},
    'холодное в/с одн':                   {'vol': 10, 'corr': 41},
    'электроснабжение день одн':          {'vol': 11, 'corr': 42},
    'электроснабжение ночь одн':          {'vol': 12, 'corr': 43},
    'электроснабжение одн':               {'vol': 13, 'corr': 44},
    # 1.2 Объём коммунальных услуг
    'водоотведение':                      {'vol': 16, 'corr': 68},
    'горячее в/с (носитель)':            {'vol': 17, 'corr': 69},
    'горячее в/с (энергия)':             {'vol': 18, 'corr': 70},
    'обращение с тко':                    {'vol': 19, 'corr': 71},
    'обращение c тко':                    {'vol': 19, 'corr': 71},
    'отопление':                          {'vol': 20, 'corr': 72},
    'холодное в/с':                       {'vol': 21, 'corr': 73},
}

# Row 90: электроэнергия ИПУ

SYSTEM_PROMPT = """Ты — эксперт по обработке российских квитанций ЖКХ.
Из текста квитанции извлеки данные ТОЛЬКО из первой основной таблицы (расчёт размера платы за жилищные и коммунальные услуги).

Верни JSON строго в таком формате:
{
  "period": "март 2026",
  "columns_present": ["перерасчеты", "задолженность", "оплачено"],
  "services": [
    {
      "name": "точное название услуги из квитанции",
      "volume": число или null,
      "perechet": число,
      "zadolzhennost": число,
      "oplacheno": число
    }
  ]
}

ПРАВИЛА ИЗВЛЕЧЕНИЯ:

Шаг 1. Определи какие колонки присутствуют в таблице:
- "Перерасчеты (доначисления +, уменьшения-)" → perechet
- "Задолженность/Переплата(-) на начало периода" → zadolzhennost
- "Оплачено, руб." → oplacheno
Запиши найденные колонки в "columns_present".

Шаг 2. Для каждой строки услуги извлеки ДОСЛОВНО из таблицы:
- "volume" = значение из колонки "Объем услуг"
- "perechet" = значение из колонки перерасчётов (0 если колонки нет)
- "zadolzhennost" = значение из колонки задолженности/переплаты (0 если колонки нет)
  ВАЖНО: сохраняй знак числа точно как в таблице. Переплата = отрицательное число (например -2 975,05).
- "oplacheno" = значение из колонки "Оплачено, руб." (0 если колонки нет)
  ВАЖНО: переписывай число точно как в таблице, не вычисляй сам.

НЕ вычисляй "correction" сам — это сделает программа.
Строку "Добровольное страхование" игнорируй.
Возвращай ТОЛЬКО валидный JSON без markdown-блоков и без комментариев."""


def extract_text_from_pdf(pdf_bytes):
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
        f.write(pdf_bytes)
        f.flush()
        reader = PdfReader(f.name)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
    os.unlink(f.name)
    return text


def normalize(name: str) -> str:
    return name.strip().lower().replace('ё', 'е')


def parse_with_claude(pdf_text: str) -> dict:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY is missing in environment")
    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=3000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": f"Текст квитанции:\n\n{pdf_text}"}]
    )
    raw = response.content[0].text.strip()
    # Strip markdown if present
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    parsed = json.loads(raw.strip())
    # Log extracted data for debugging
    print(f"\n=== AI извлёк данные ===")
    print(f"Период: {parsed.get('period')}")
    print(f"Колонки: {parsed.get('columns_present')}")
    for s in parsed.get('services', []):
        print(f"  {s['name']}: vol={s.get('volume')}, perechet={s.get('perechet',0)}, zadolzh={s.get('zadolzhennost',0)}, oplach={s.get('oplacheno',0)}, corr={s.get('correction',0)}")
    print("========================\n")
    return parsed


def find_month_column(wb, period: str) -> int | None:
    """Find first empty column by checking row 5 (real volume data, not formula headers).
    Row 2 has EOMONTH formulas which appear as None in read_only mode — unreliable.
    Row 5 = Взнос на капитальный ремонт volumes — reliable indicator of filled months.
    Jan=col5(E), Feb=col6(F), Mar=col7(G), Apr=col8(H), etc.
    """
    ws = wb['Рассчёты']
    for col in range(5, 18):
        val = ws.cell(row=5, column=col).value
        if val is None or val == '':
            return col
    return None


def calc_correction(svc: dict, columns_present: list) -> float:
    """
    Формула из листа Описание:
    Перерасчёты + Задолженность/Переплата(-) - Оплачено
    
    Ключевое правило: Задолженность и Оплачено нивелируют друг друга когда равны.
    Например: Задолженность=1412.40, Оплачено=1412.40 → вклад = 1412.40 - 1412.40 = 0
    Например: Задолженность=-2975.05, Оплачено=0 → вклад = -2975.05 - 0 = -2975.05
    """
    perechet = float(svc.get('perechet') or 0)
    zadolzh  = float(svc.get('zadolzhennost') or 0)
    oplach   = float(svc.get('oplacheno') or 0)
    result = perechet + zadolzh - oplach
    # Round to 2 decimal places to avoid float noise
    return round(result, 2)


def write_formulas(ws, col: int):
    """Write all Excel formulas for the given column.
    Formulas are adapted from the original template (col E = col 5).
    Pattern: replace 'E' with the target column letter in all formulas.
    """
    from openpyxl.utils import get_column_letter
    C = get_column_letter(col)  # e.g. 'H' for col 8

    def f(formula: str) -> str:
        """Replace column letter E with target column in formula."""
        # Replace cell refs like E5, E25, E47:E55 etc (E followed by digit)
        import re
        return re.sub(r'\bE(\d)', lambda m: f'{C}{m.group(1)}', formula)

    # 2.1 Начислено по тарифу (жилищные): rows 25-32
    # Pattern: =$D{vol_row} * {C}{vol_row}
    vol_rows_zhil = {25: 5, 26: 6, 27: 7, 28: 8, 29: 9, 30: 10, 31: 11, 32: 12}
    for calc_row, vol_row in vol_rows_zhil.items():
        ws.cell(row=calc_row, column=col).value = f'=$D{vol_row} * {C}{vol_row}'
    # Row 33: Электроснабжение ОДН — значение берётся из корректировок, формулы нет
    # Row 34: Сумма начислений жилищных
    ws.cell(row=34, column=col).value = f'=SUM({C}25:{C}33)'

    # 2.2 Корректировки (жилищные): rows 36-44 — заполняются данными из PDF, не формулы
    # Row 45: Сумма корректировок жилищных
    ws.cell(row=45, column=col).value = f'=SUM({C}36:{C}44)'

    # 2.3 К оплате (жилищные): rows 47-54
    pairs_zhil = [(47,25,36),(48,26,37),(49,27,38),(50,28,39),
                  (51,29,40),(52,30,41),(53,31,42),(54,32,43)]
    for pay_row, calc_row, corr_row in pairs_zhil:
        ws.cell(row=pay_row, column=col).value = f'=SUM({C}{calc_row},{C}{corr_row})'
    # Row 55: Электроснабжение ОДН к оплате
    ws.cell(row=55, column=col).value = f'=SUM({C}33,{C}44)'
    # Row 56: Итого к оплате жилищные
    ws.cell(row=56, column=col).value = f'=SUM({C}47:{C}55)'

    # 3.1 Начислено по тарифу (коммунальные): rows 60-65
    vol_rows_kom = {60: 16, 61: 17, 62: 18, 63: 19, 64: 20, 65: 21}
    for calc_row, vol_row in vol_rows_kom.items():
        ws.cell(row=calc_row, column=col).value = f'=$D{vol_row} * {C}{vol_row}'
    # Row 66: Сумма начислений коммунальные
    ws.cell(row=66, column=col).value = f'=SUM({C}60:{C}65)'

    # 3.2 Корректировки (коммунальные): rows 68-73 — данные из PDF
    # Row 74: Сумма корректировок коммунальные
    ws.cell(row=74, column=col).value = f'=SUM({C}68:{C}73)'

    # 3.3 К оплате (коммунальные): rows 76-81, с IF > 0
    pairs_kom = [(76,60,68),(77,61,69),(78,62,70),(79,63,71),(80,64,72),(81,65,73)]
    for pay_row, calc_row, corr_row in pairs_kom:
        ws.cell(row=pay_row, column=col).value = f'=IF(SUM({C}{calc_row},{C}{corr_row})>0,SUM({C}{calc_row},{C}{corr_row}),0)'
    # Row 82: Итого к оплате коммунальные
    ws.cell(row=82, column=col).value = f'=SUM({C}76:{C}81)'

    # 4. Всего к оплате по квитанции
    ws.cell(row=84, column=col).value = f'=SUM({C}56,{C}82)'
    # 4.1 Арендатор
    ws.cell(row=86, column=col).value = f'=SUMIFS({C}$23:{C}$82,$B$23:$B$82,$B86)'
    # 4.2 Арендодатель
    ws.cell(row=87, column=col).value = f'=SUMIFS({C}$23:{C}$82,$B$23:$B$82,$B87)'
    # Check
    ws.cell(row=88, column=col).value = f'=SUM({C}86:{C}87)-{C}84'
    # 6. Всего к оплате Арендатором
    ws.cell(row=92, column=col).value = f'=SUM({C}86,{C}90)'


def write_to_excel(parsed: dict, electricity_ipu: float, output_path: str, source_xlsx: str = None):
    src = source_xlsx if source_xlsx else XLSX_DEFAULT
    shutil.copy(src, output_path)
    wb = load_workbook(output_path)
    ws = wb['Рассчёты']

    col = find_month_column(wb, parsed.get('period', ''))
    if col is None:
        raise ValueError("Нет свободных колонок в таблице (все 12 месяцев заполнены)")

    # Always write formulas first — works even if template has no data
    write_formulas(ws, col)

    columns_present = parsed.get('columns_present', [])
    services = parsed.get('services', [])

    for svc in services:
        key = normalize(svc['name'])
        row_info = ROW_MAP.get(key)
        if not row_info:
            for k, v in ROW_MAP.items():
                if k in key or key in k:
                    row_info = v
                    break

        if row_info:
            vol = svc.get('volume')
            corr = calc_correction(svc, columns_present)

            if vol is not None:
                ws.cell(row=row_info['vol'], column=col).value = vol
            ws.cell(row=row_info['corr'], column=col).value = corr

            print(f"  Запись [{svc['name']}]: vol={vol}, perechet={svc.get('perechet',0)}, zadolzh={svc.get('zadolzhennost',0)}, oplach={svc.get('oplacheno',0)} → corr={corr}")

    # Электроэнергия ИПУ — row 90
    ws.cell(row=90, column=col).value = electricity_ipu

    wb.save(output_path)
    return col


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process():
    if 'pdf' not in request.files:
        return jsonify({'error': 'PDF файл не загружен'}), 400

    pdf_file = request.files['pdf']
    electricity = request.form.get('electricity', '0')
    xlsx_file = request.files.get('xlsx')  # optional upload

    try:
        electricity_val = float(electricity.replace(',', '.'))
    except ValueError:
        return jsonify({'error': 'Некорректная сумма электроэнергии'}), 400

    # Save uploaded xlsx to temp if provided
    source_xlsx = None
    if xlsx_file and xlsx_file.filename:
        tmp_xlsx = os.path.join(tempfile.gettempdir(), 'source_kvitanciya.xlsx')
        xlsx_file.save(tmp_xlsx)
        source_xlsx = tmp_xlsx

    pdf_bytes = pdf_file.read()

    try:
        pdf_text = extract_text_from_pdf(pdf_bytes)
    except Exception as e:
        return jsonify({'error': f'Ошибка чтения PDF: {str(e)}'}), 400

    try:
        parsed = parse_with_claude(pdf_text)
    except Exception as e:
        return jsonify({'error': f'Ошибка AI-обработки: {str(e)}'}), 500

    # Output filename matches period from receipt
    period = parsed.get('period', 'updated').replace(' ', '_')
    out_name = f'Квитанции_{period}.xlsx'
    output_path = os.path.join(tempfile.gettempdir(), out_name)

    try:
        col = write_to_excel(parsed, electricity_val, output_path, source_xlsx=source_xlsx)
    except Exception as e:
        return jsonify({'error': f'Ошибка записи в Excel: {str(e)}'}), 500

    # Return parsed data + success
    return jsonify({
        'success': True,
        'period': parsed.get('period'),
        'services_count': len(parsed.get('services', [])),
        'electricity_ipu': electricity_val,
        'services': parsed.get('services', []),
        'download_ready': True,
        'filename': out_name
    })


@app.route('/download')
def download():
    fname = request.args.get('file', 'Квитанции_updated.xlsx')
    output_path = os.path.join(tempfile.gettempdir(), fname)
    if not os.path.exists(output_path):
        return jsonify({'error': 'Файл не найден, сначала обработайте квитанцию'}), 404
    return send_file(output_path, as_attachment=True, download_name=fname)

@app.route('/debug')
def debug():
    return jsonify({
        "key_exists": "ANTHROPIC_API_KEY" in os.environ,
        "key_preview": (os.environ.get("ANTHROPIC_API_KEY") or "")[:6] + "..." if os.environ.get("ANTHROPIC_API_KEY") else None
    })

@app.route('/debug-full')
def debug_full():
    import os
    return {
        "env_keys": list(os.environ.keys())
    }

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    app.run(host='0.0.0.0', port=port)
